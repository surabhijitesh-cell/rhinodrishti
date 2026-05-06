"""Document upload, URL analysis, and contextual intelligence assessment."""
from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
import io
import os
import re
import uuid
import json
import anthropic
from shared import db, uploads_col, intelligence_col, patterns_col, logger
from llm_client import get_client, MODEL

router = APIRouter()


# ============================================================
# Robust JSON extraction helper
# ============================================================

def _extract_json(text: str) -> dict:
    """
    Extract a JSON object from Claude's response text.

    Handles:
    - Markdown code fences (```json ... ```)
    - Trailing commas before } or ] (common Claude quirk)
    - Extra prose before/after the JSON block
    - Truncated responses (graceful fallback)
    """
    # 1. Strip markdown code fences
    cleaned = re.sub(r"```(?:json)?\s*", "", text).strip()
    cleaned = re.sub(r"```\s*$", "", cleaned).strip()

    # 2. Try parsing the whole cleaned string directly
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # 3. Find the outermost {...} block
    start = cleaned.find("{")
    if start != -1:
        # Walk backwards from end to find matching closing brace
        depth = 0
        end = -1
        for i, ch in enumerate(cleaned[start:], start):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        if end != -1:
            json_str = cleaned[start : end + 1]
            try:
                return json.loads(json_str)
            except json.JSONDecodeError:
                # 4. Fix trailing commas before } or ] (common Claude formatting issue)
                fixed = re.sub(r",\s*([}\]])", r"\1", json_str)
                try:
                    return json.loads(fixed)
                except json.JSONDecodeError:
                    pass

    # 5. Final fallback — return whatever we got as a plain summary
    logger.warning("_extract_json: could not parse JSON from LLM response")
    return {
        "executive_summary": text[:600],
        "error": "Could not parse structured response — raw text stored",
    }


class URLAnalysisRequest(BaseModel):
    url: str
    analysis_query: str = ""


class TextAnalysisRequest(BaseModel):
    text: str
    analysis_query: str = ""


# ============================================================
# Fetch current security context for AI prompts
# ============================================================

async def _get_security_context() -> str:
    """Build a snapshot of the current security environment from recent intelligence."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()

    # Recent high-priority items
    recent_items = await intelligence_col.find(
        {"processed": True, "severity": {"$in": ["critical", "high"]}, "published_at": {"$gte": cutoff}},
        {"_id": 0, "title": 1, "state": 1, "severity": 1, "tags": 1, "ai_summary": 1}
    ).sort("priority_score", -1).limit(15).to_list(15)

    # Active patterns
    recent_patterns = await patterns_col.find(
        {"detected_at": {"$gte": cutoff}},
        {"_id": 0, "pattern_type": 1, "description": 1, "severity": 1, "states": 1}
    ).sort("detected_at", -1).limit(10).to_list(10)

    # Build context string
    ctx = "=== CURRENT NER SECURITY ENVIRONMENT (Last 7 Days) ===\n\n"
    ctx += "--- HIGH PRIORITY DEVELOPMENTS ---\n"
    for i, item in enumerate(recent_items, 1):
        ctx += f"{i}. [{item.get('state','')}|{item.get('severity','').upper()}] {item.get('title','')}\n"
        if item.get('ai_summary'):
            ctx += f"   Summary: {item['ai_summary'][:150]}\n"
        tags = item.get('tags', [])
        if tags:
            ctx += f"   Tags: {', '.join(tags[:5])}\n"

    ctx += "\n--- DETECTED PATTERNS ---\n"
    for p in recent_patterns:
        ctx += f"- [{p.get('severity','').upper()}] {p.get('pattern_type','')}: {p.get('description','')[:120]}\n"
        states = p.get('states', [])
        if states:
            ctx += f"  Affected: {', '.join(states)}\n"

    return ctx


CONTEXTUAL_ANALYSIS_PROMPT = """You are a SENIOR MILITARY INTELLIGENCE ANALYST at India's Northeast Regional Command.

You have been provided:
1. A document/article/report submitted for analysis
2. The CURRENT SECURITY ENVIRONMENT — a snapshot of the latest 7 days of intelligence from the region

Your task: Provide a COMPREHENSIVE CONTEXTUAL INTELLIGENCE ASSESSMENT of the submitted document.

Structure your response STRICTLY as JSON:
{
  "executive_summary": "3-4 sentence overview of what this document reports and why it matters",
  "threat_classification": {
    "severity": "CRITICAL / HIGH / MEDIUM / LOW",
    "threat_category": "One of: INSURGENCY, BORDER SECURITY, DRUG TRAFFICKING, ARMS SMUGGLING, ETHNIC CONFLICT, POLITICAL INSTABILITY, CROSS-BORDER INFILTRATION, TERRORISM, CIVIL UNREST, NATURAL DISASTER, OTHER",
    "confidence": "HIGH / MEDIUM / LOW"
  },
  "pattern_analysis": {
    "matches_existing_pattern": true/false,
    "pattern_description": "How this fits or diverges from detected patterns in the region. Be specific.",
    "escalation_indicator": "ESCALATING / STABLE / DE-ESCALATING / NEW_THREAT"
  },
  "relevance_assessment": {
    "relevance_score": 1-10,
    "primary_region": "State or country most affected",
    "secondary_regions": ["other affected areas"],
    "relevance_explanation": "How this connects to the current NER security scenario"
  },
  "key_entities": {
    "actors": ["Named groups, forces, individuals"],
    "locations": ["Specific places mentioned"],
    "events": ["Key events described"]
  },
  "recommended_actions": [
    {"priority": "IMMEDIATE/HIGH/MEDIUM/LOW", "action": "Specific actionable recommendation"},
    ...
  ],
  "intelligence_gaps": ["What information is missing that would improve assessment"],
  "cross_references": "Which of the current security developments (from the context provided) are most related to this document and how"
}

RULES:
- Be specific, not generic. Reference actual locations, actors, events.
- recommended_actions must be ACTIONABLE — not vague suggestions.
- If the document is NOT relevant to NER security, say so clearly but still provide the classification.
- Use the current security context to make connections the raw document alone would not reveal."""


# ============================================================
# Endpoints
# ============================================================

@router.get("/uploaded-documents")
async def get_uploaded_documents():
    docs = await uploads_col.find({}, {"_id": 0}).sort("uploaded_at", -1).to_list(100)
    return {"documents": docs, "count": len(docs)}


@router.post("/upload-document")
async def upload_document(file: UploadFile = File(...), background_tasks: BackgroundTasks = None):
    allowed_types = {
        "application/pdf": "pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
        "application/msword": "doc",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/vnd.ms-excel": "xls",
        "text/plain": "txt"
    }
    content_type = file.content_type
    if content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="File type not supported. Allowed: PDF, Word, Excel, TXT")

    file_type = allowed_types[content_type]
    file_content = await file.read()

    extracted_text = ""
    try:
        if file_type == "pdf":
            from PyPDF2 import PdfReader
            pdf_reader = PdfReader(io.BytesIO(file_content))
            for page in pdf_reader.pages:
                extracted_text += page.extract_text() or ""
        elif file_type in ["docx", "doc"]:
            from docx import Document
            doc = Document(io.BytesIO(file_content))
            extracted_text = "\n".join([p.text for p in doc.paragraphs])
        elif file_type in ["xlsx", "xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content))
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    extracted_text += " | ".join([str(c) for c in row if c]) + "\n"
        elif file_type == "txt":
            extracted_text = file_content.decode('utf-8', errors='ignore')
    except Exception as e:
        logger.error(f"Error extracting text from {file.filename}: {e}")
        extracted_text = f"Error extracting text: {str(e)}"

    doc_record = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "file_type": file_type,
        "source_type": "file",
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "extracted_text": extracted_text[:15000],
        "analysis": None,
        "processed": False,
    }
    await uploads_col.insert_one(doc_record)

    if background_tasks:
        background_tasks.add_task(_run_contextual_analysis, doc_record["id"])

    return {"message": "Document uploaded — analysis starting", "document_id": doc_record["id"], "filename": file.filename}


@router.post("/analyze-url")
async def analyze_url(data: URLAnalysisRequest, background_tasks: BackgroundTasks = None):
    """Fetch a URL and run contextual intelligence analysis."""
    url = data.url.strip()
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="Invalid URL")

    # Scrape the URL
    try:
        import httpx
        from bs4 import BeautifulSoup
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            resp = await client.get(url, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
        title = soup.title.string if soup.title else url
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to fetch URL: {str(e)}")

    doc_record = {
        "id": str(uuid.uuid4()),
        "filename": title[:200] if title else url[:200],
        "file_type": "url",
        "source_type": "url",
        "source_url": url,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "extracted_text": text[:15000],
        "analysis_query": data.analysis_query,
        "analysis": None,
        "processed": False,
    }
    await uploads_col.insert_one(doc_record)

    if background_tasks:
        background_tasks.add_task(_run_contextual_analysis, doc_record["id"])

    return {"message": "URL fetched — analysis starting", "document_id": doc_record["id"], "filename": doc_record["filename"]}


@router.get("/uploaded-documents/{doc_id}")
async def get_document_detail(doc_id: str):
    doc = await uploads_col.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc


@router.delete("/uploaded-documents/{doc_id}")
async def delete_uploaded_document(doc_id: str):
    result = await uploads_col.delete_one({"id": doc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted"}


# ============================================================
# Add URL to Intelligence Feed
# ============================================================

class AddToFeedRequest(BaseModel):
    url: str
    title: str = ""
    severity: str = "medium"
    priority_score: int = 50
    threat_category: str = ""
    state: str = ""
    ai_summary: str = ""
    is_cross_border: bool = False
    tags: list = []


@router.post("/add-to-feed")
async def add_url_to_feed(data: AddToFeedRequest):
    """Scrape a URL and add it directly to the intelligence feed with user-defined parameters."""
    url = data.url.strip()
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="Invalid URL")

    # Check for duplicate URL
    existing = await intelligence_col.find_one({"source_url": url}, {"_id": 0, "id": 1})
    if existing:
        raise HTTPException(status_code=409, detail="This URL already exists in the intelligence feed")

    # Scrape the URL for content (best-effort — still add even if scraping fails)
    raw_content = ""
    scraped_title = ""
    try:
        import httpx
        from bs4 import BeautifulSoup
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
            resp = await client.get(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"})
            resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        raw_content = soup.get_text(separator="\n", strip=True)[:5000]
        scraped_title = soup.title.string.strip() if soup.title and soup.title.string else ""
    except Exception as e:
        logger.warning(f"URL scrape failed for {url}: {e} — adding with user-provided data")
        if not data.title.strip():
            raise HTTPException(status_code=400, detail=f"Could not fetch URL and no title provided. Error: {str(e)}")

    title = data.title.strip() or scraped_title or url[:200]
    severity = data.severity.lower() if data.severity else "medium"
    if severity not in ("critical", "high", "medium", "low"):
        severity = "medium"

    priority_score = max(0, min(100, data.priority_score))

    # Translate non-English title and content
    from shared import has_non_latin_chars, translate_to_english
    summary = data.ai_summary or title
    if has_non_latin_chars(title):
        translated_title = await translate_to_english(title)
        if translated_title and translated_title != title:
            title = translated_title
    if has_non_latin_chars(summary):
        translated_summary = await translate_to_english(summary)
        if translated_summary and translated_summary != summary:
            summary = translated_summary
    if raw_content and has_non_latin_chars(raw_content[:500]):
        translated_content = await translate_to_english(raw_content[:3000])
        if translated_content:
            raw_content = translated_content

    item = {
        "id": str(uuid.uuid4()),
        "title": title,
        "source": "Manual Upload",
        "source_url": url,
        "published_at": datetime.now(timezone.utc).isoformat(),
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "raw_content": raw_content,
        "severity": severity,
        "priority_score": priority_score,
        "threat_category": data.threat_category or "Unclassified",
        "state": data.state or "Unknown",
        "ai_summary": summary,
        "is_cross_border": data.is_cross_border,
        "tags": data.tags or [],
        "countries_involved": [],
        "processed": True,
        "is_relevant": True,
        "why_it_matters": summary,
        "attention_level": "Priority Monitoring" if priority_score >= 60 else "Standard Monitoring",
        "potential_impact": "",
    }

    await intelligence_col.insert_one(item)
    logger.info(f"Manual feed item added: '{title[:60]}' severity={severity} priority={priority_score}")

    return {
        "message": "Article added to intelligence feed",
        "item_id": item["id"],
        "title": title,
        "severity": severity,
        "priority_score": priority_score,
    }


# ============================================================
# Contextual AI Analysis
# ============================================================

async def _run_contextual_analysis(doc_id: str):
    """Run comprehensive contextual intelligence analysis on a document."""
    doc = await uploads_col.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        return

    text = doc.get("extracted_text", "")
    if not text or len(text) < 20:
        await uploads_col.update_one(
            {"id": doc_id},
            {"$set": {"processed": True, "analysis": {"error": "No extractable text found"}}}
        )
        return

    try:
        security_context = await _get_security_context()
        extra_query = doc.get("analysis_query", "")

        # Cap document text at 4 000 chars to keep prompt size manageable and
        # reduce the chance of the LLM response being truncated at max_tokens.
        user_prompt = f"=== DOCUMENT/ARTICLE SUBMITTED FOR ANALYSIS ===\n\n{text[:4000]}\n\n"
        user_prompt += f"=== SECURITY CONTEXT ===\n{security_context}\n\n"
        if extra_query:
            user_prompt += f"=== SPECIFIC ANALYSIS REQUEST ===\n{extra_query}\n\n"
        user_prompt += "Provide your COMPREHENSIVE CONTEXTUAL INTELLIGENCE ASSESSMENT as JSON."

        client = get_client()
        # max_tokens raised to 3000 so the full JSON body is never truncated.
        # timeout=90 gives Render-hosted background tasks a generous window
        # while still surfacing a clean error if the API is unresponsive.
        response = await client.messages.create(
            model=MODEL,
            max_tokens=3000,
            timeout=90.0,
            system=[
                {
                    "type": "text",
                    "text": CONTEXTUAL_ANALYSIS_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            messages=[{"role": "user", "content": user_prompt}],
            extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
        )
        response_text = response.content[0].text

        # Robust JSON extraction — handles markdown fences, trailing commas,
        # and partial responses much better than a plain regex+json.loads.
        analysis = _extract_json(response_text)

        await uploads_col.update_one(
            {"id": doc_id},
            {"$set": {
                "analysis": analysis,
                "processed": True,
                "region": analysis.get("relevance_assessment", {}).get("primary_region", ""),
            }}
        )
        logger.info(
            f"Contextual analysis complete for {doc_id}: "
            f"{analysis.get('threat_classification', {}).get('severity', '?')}"
        )

    except (anthropic.APITimeoutError, anthropic.APIConnectionError) as e:
        msg = (
            "Analysis request timed out. The document may be too large or the AI "
            "service is temporarily slow. Please try again in a few minutes."
        )
        logger.error(f"Analysis timeout/connection error for {doc_id}: {e}")
        await uploads_col.update_one(
            {"id": doc_id},
            {"$set": {"processed": True, "analysis": {"error": msg}}}
        )

    except Exception as e:
        logger.error(f"Analysis failed for {doc_id}: {e}")
        await uploads_col.update_one(
            {"id": doc_id},
            {"$set": {"processed": True, "analysis": {"error": str(e)}}}
        )
